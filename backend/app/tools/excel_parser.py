"""
File parser — sends uploaded Excel / PDF files to Claude Opus 4.6
for structured portfolio data extraction.

Single entry point:
- parse_file_with_claude(file_path, anthropic_client, client_name, client_type)

Both Excel and PDF files are converted to text snapshots before being sent
to Claude, keeping token usage manageable.
"""

from __future__ import annotations

import json
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from pypdf import PdfReader

from app.agents.base import MODEL

logger = logging.getLogger(__name__)


# ======================================================================
# Text snapshot for Excel files (Claude cannot read binary .xlsx)
# ======================================================================

def _format_cell_value(val: Any) -> str:
    """Format a cell value for the text snapshot, keeping it readable."""
    if val is None:
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return f"{val:.4f}".rstrip("0").rstrip(".")
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def extract_excel_snapshot(
    file_path: str, max_rows: int = 200, max_cols: int = 70
) -> str:
    """
    Read an Excel file and produce a text snapshot of its contents
    suitable for sending to Claude for interpretation.

    Returns a string with sheet names, dimensions, and a TSV-like text
    table of cells per sheet.
    """
    MAX_SNAPSHOT_CHARS = 40_000
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot open Excel file: {exc}") from exc

    parts: list[str] = []
    parts.append(f"Workbook: {Path(file_path).name}")
    parts.append(f"Sheets: {', '.join(wb.sheetnames)}")
    parts.append("")

    total_len = sum(len(p) for p in parts)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = f"=== Sheet: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ==="
        parts.append(header)
        total_len += len(header) + 1

        rows_to_read = min(max_rows, ws.max_row or 0)
        cols_to_read = min(max_cols, ws.max_column or 0)

        for r in range(1, rows_to_read + 1):
            cells: list[str] = []
            for c in range(1, cols_to_read + 1):
                try:
                    cell = ws.cell(row=r, column=c)
                    cells.append(_format_cell_value(cell.value))
                except Exception:
                    cells.append("")
            line = "\t".join(cells)
            # Skip completely empty rows
            if not line.replace("\t", "").strip():
                continue
            total_len += len(line) + 1
            if total_len > MAX_SNAPSHOT_CHARS:
                parts.append("... [snapshot truncated] ...")
                wb.close()
                return "\n".join(parts)
            parts.append(line)

        parts.append("")  # blank line between sheets

    wb.close()
    return "\n".join(parts)


# ======================================================================
# Text snapshot for PDF files
# ======================================================================

def extract_pdf_snapshot(file_path: str, max_chars: int = 40_000) -> str:
    """
    Extract text from a PDF and produce a snapshot suitable for Claude.

    Returns a string with page-by-page text content, capped at *max_chars*.
    """
    try:
        reader = PdfReader(file_path)
    except Exception as exc:
        raise ValueError(f"Cannot open PDF file: {exc}") from exc

    parts: list[str] = []
    filename = Path(file_path).name
    num_pages = len(reader.pages)
    parts.append(f"PDF: {filename} ({num_pages} pages)")
    parts.append("")

    total_len = sum(len(p) for p in parts)

    for i, page in enumerate(reader.pages):
        try:
            page_text = page.extract_text() or ""
        except Exception:
            page_text = ""

        if not page_text.strip():
            continue

        header = f"=== Page {i + 1} ==="
        total_len += len(header) + len(page_text) + 2
        if total_len > max_chars:
            parts.append("... [snapshot truncated] ...")
            break
        parts.append(header)
        parts.append(page_text)
        parts.append("")

    return "\n".join(parts)


# ======================================================================
# System prompt for Claude-based file parsing
# ======================================================================

_PARSE_SYSTEM_PROMPT = """\
You are a financial data extraction assistant for a captive insurance asset \
management firm. You receive either a text snapshot of an Excel workbook or a \
PDF document containing portfolio data.

Your job is to extract ALL meaningful financial data and map it into the JSON \
schema below.

CLEARWATER ANALYTICS EXPORT PATTERNS:
The most common input is a Clearwater Analytics export. Recognise these patterns:

Board Reports sheet:
- "Net Performance" section: periods (QTD, YTD, 1Y, 3Y, 5Y, 10Y, Inception) \
with Total Return and Assigned Index Return.
- "Allocation vs Guidelines" section: strategies with market value, target %, \
actual %, unrealized gain/loss.
- "Market Value History" section: period begin/end dates with base market value.
- "Portfolio Rollforward" section: labelled rows (beginning MV, contributions, \
withdrawals, investment income, realised/unrealised gains, fees, ending MV) \
with dollar values.
- "Contribution by Strategy" section: strategy names with contribution values.

SMA Holdings sheet:
- Individual bond rows with: identifier, issuer/description, sector, \
sector_category, duration, yield, S&P rating, Moody's rating, maturity date, \
market value, % of portfolio.
- A summary/total row (marked with "---" in identifier) containing weighted \
averages for duration, yield, ratings, maturity, and total market value.

Returns are raw decimals (0.097 = 9.7%). "---" means data unavailable.

GENERIC FINANCIAL DOCUMENTS:
If the document is not a Clearwater export, creatively map whatever financial \
data you find:
- "performance": any time-series metrics. Use "total_return" for the PRIMARY \
numeric metric (e.g. revenue, earnings, NAV, price — whatever the key figure \
is). Use "index_return" for a comparison/benchmark if available. NEVER leave \
total_return as null if there is a numeric value for that period — the slide \
builder will skip rows where both total_return and index_return are null.
- "allocation": any breakdown/composition data (by segment, category, strategy).
- "roll_forward": any waterfall/bridge data (cash flow, P&L line items).
- "contributions": any strategy contribution data.
- "mv_history": any market value over time data.
- "sma_holdings": individual security/bond holdings.
- "sma_summary": aggregated portfolio statistics.

IMPORTANT: Extract ALL data you can find — every section, every table, every \
metric. The system will automatically generate slides for whatever data is present. \
Do not skip any section.

TARGET OUTPUT SCHEMA:
{
  "performance": [
    {"period": "<label>", "total_return": <float|null>, "index_return": <float|null>}
  ],
  "allocation": [
    {"strategy": "<name>", "market_value": <float>, "actual_pct": <float|null>, \
"target_pct": <float|null>, "unrealized_gain_loss": <float|null>}
  ],
  "roll_forward": [
    {"label": "<item>", "value": <float|null>}
  ],
  "contributions": [
    {"strategy": "<name>", "contribution": <float|null>}
  ],
  "mv_history": [
    {"period_begin": "<ISO date>", "period_end": "<ISO date>", "market_value": <float>}
  ],
  "sma_holdings": [
    {"identifier": "<id>", "issuer": "<name>", "sector": "<sector>", \
"sector_category": "<cat|null>", "duration": <float|null>, \
"yield_to_worst": <float|null>, "sp_rating": "<rating|null>", \
"moody_rating": "<rating|null>", "maturity_date": "<ISO date|null>", \
"market_value": <float|null>, "pct_of_portfolio": <float|null>}
  ],
  "sma_summary": {
    "total_market_value": <float>, "avg_duration": <float>, "avg_yield": <float>, \
"num_holdings": <int>, "sector_allocation": {"<sector>": <pct>}, \
"credit_quality": {"<rating>": <pct>}, "avg_maturity_date": "<ISO date|null>", \
"weighted_sp_rating": "<rating|null>", "weighted_moody_rating": "<rating|null>"
  }
}

RULES:
- Populate every section you can find data for. NEVER return all empty arrays.
- You MUST populate at least "performance" or "allocation".
- Values should be raw numbers (no currency symbols, no commas).
- Percentages as decimals for actual_pct, target_pct, pct_of_portfolio (0.05 = 5%).
- Dates as ISO format strings (YYYY-MM-DD).
- Return ONLY the JSON object — no markdown fences, no commentary.
- KEEP OUTPUT COMPACT: no extra whitespace, no indentation.
- For sma_holdings: return ONLY the sma_summary with aggregated statistics. \
Set sma_holdings to an empty array — the system will extract individual \
holdings separately. Focus on computing accurate summary stats: \
total_market_value, avg_duration, avg_yield, num_holdings, sector_allocation, \
credit_quality, weighted ratings.
"""


def _extract_holdings_from_snapshot(file_path: str) -> list[dict]:
    """Extract individual holdings rows from an Excel snapshot using openpyxl.

    This is used for SMA Holdings files where the data is structured tabular
    data with clear column headers — no AI needed.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return []

    # Find SMA Holdings sheet
    holdings_sheet = None
    for name in wb.sheetnames:
        if "holdings" in name.lower() or "sma" in name.lower():
            holdings_sheet = wb[name]
            break
    if not holdings_sheet:
        wb.close()
        return []

    # Find header row (look for "Identifier" or "Description" in first 10 rows)
    header_row = None
    col_map: dict[str, int] = {}
    target_headers = {
        "identifier": "identifier", "description": "issuer", "sector": "sector",
        "performa sector 2": "sector", "performa sector category": "sector_category",
        "sector category": "sector_category",
        "duration": "duration", "yield": "yield_to_worst",
        "s&p rating": "sp_rating", "s&p": "sp_rating",
        "moody": "moody_rating", "moody's rating": "moody_rating",
        "final maturity": "maturity_date", "maturity": "maturity_date",
        "base market value": "market_value", "market value": "market_value",
        "% of market value": "pct_of_portfolio", "% of portfolio": "pct_of_portfolio",
    }

    for r in range(1, min(15, (holdings_sheet.max_row or 0) + 1)):
        for c in range(1, min(20, (holdings_sheet.max_column or 0) + 1)):
            val = holdings_sheet.cell(row=r, column=c).value
            if val is None:
                continue
            cell_text = str(val).strip().lower()
            for pattern, field in target_headers.items():
                if pattern in cell_text and field not in col_map:
                    col_map[field] = c
                    header_row = r

    if not header_row or not col_map:
        wb.close()
        return []

    # Extract data rows
    from datetime import datetime as _dt
    holdings: list[dict] = []
    data_start = header_row + 1

    for r in range(data_start, (holdings_sheet.max_row or 0) + 1):
        # Check if row has issuer data
        issuer_col = col_map.get("issuer")
        if issuer_col:
            issuer_val = holdings_sheet.cell(row=r, column=issuer_col).value
            if not issuer_val or str(issuer_val).strip() in ("", "---"):
                # Could be summary row or empty — check identifier
                id_col = col_map.get("identifier")
                if id_col:
                    id_val = holdings_sheet.cell(row=r, column=id_col).value
                    if str(id_val).strip() == "---":
                        continue  # Skip summary row
                continue

        def _get_float(col_name: str) -> float | None:
            c = col_map.get(col_name)
            if not c:
                return None
            v = holdings_sheet.cell(row=r, column=c).value
            if v is None or str(v).strip() in ("", "---", "NA", "N/A"):
                return None
            try:
                return float(str(v).replace(",", "").replace("$", ""))
            except (ValueError, TypeError):
                return None

        def _get_str(col_name: str) -> str | None:
            c = col_map.get(col_name)
            if not c:
                return None
            v = holdings_sheet.cell(row=r, column=c).value
            if v is None:
                return None
            s = str(v).strip()
            return s if s and s not in ("---", "NA", "N/A") else None

        def _get_date(col_name: str) -> str | None:
            c = col_map.get(col_name)
            if not c:
                return None
            v = holdings_sheet.cell(row=r, column=c).value
            if isinstance(v, _dt):
                return v.strftime("%Y-%m-%d")
            if v is None:
                return None
            s = str(v).strip()
            return s if s and s not in ("---", "") else None

        holding = {
            "identifier": _get_str("identifier"),
            "issuer": _get_str("issuer"),
            "sector": _get_str("sector") or "Other",
            "sector_category": _get_str("sector_category"),
            "duration": _get_float("duration"),
            "yield_to_worst": _get_float("yield_to_worst"),
            "sp_rating": _get_str("sp_rating"),
            "moody_rating": _get_str("moody_rating"),
            "maturity_date": _get_date("maturity_date"),
            "market_value": _get_float("market_value"),
            "pct_of_portfolio": _get_float("pct_of_portfolio"),
        }
        if holding.get("issuer"):
            holdings.append(holding)

    wb.close()
    return holdings


def _repair_truncated_json(text: str) -> dict | None:
    """Attempt to repair JSON that was truncated mid-generation (max_tokens hit)."""
    # Find the opening brace
    start = text.find("{")
    if start == -1:
        return None
    truncated = text[start:]

    # Try progressively closing open brackets/braces from the truncation point
    # First, remove the last incomplete element (likely a partial object)
    for trim in range(min(200, len(truncated)), 0, -1):
        candidate = truncated[:len(truncated) - trim]
        # Count open/close brackets
        opens = candidate.count("[") - candidate.count("]")
        braces = candidate.count("{") - candidate.count("}")
        # Close them
        suffix = "]" * opens + "}" * braces
        try:
            return json.loads(candidate + suffix)
        except json.JSONDecodeError:
            continue
    return None


# ======================================================================
# Main entry point — parse any file with Claude
# ======================================================================

def parse_file_with_claude(
    file_path: str,
    anthropic_client: Any,
    client_name: str = "Client",
) -> dict:
    """
    Parse an uploaded Excel or PDF file using Claude.

    - Excel (.xlsx / .xls): converted to a text snapshot then sent as text.
    - PDF (.pdf): text extracted via pypdf, sent as text.

    Returns a dict matching the PortfolioData schema.
    """
    ext = Path(file_path).suffix.lower()
    filename = Path(file_path).name

    # Build the user message content blocks
    if ext in (".xlsx", ".xls"):
        snapshot = extract_excel_snapshot(file_path)
        content: list[dict[str, Any]] = [
            {
                "type": "text",
                "text": (
                    f"Client name: {client_name}\n"
                    f"File: {filename}\n\n"
                    f"Below is a text snapshot of the uploaded Excel workbook. "
                    f"Extract ALL data and return JSON matching "
                    f"the target schema.\n\n"
                    f"{snapshot}"
                ),
            }
        ]
    elif ext == ".pdf":
        snapshot = extract_pdf_snapshot(file_path)
        content = [
            {
                "type": "text",
                "text": (
                    f"Client name: {client_name}\n"
                    f"File: {filename}\n\n"
                    f"Below is the extracted text from the uploaded PDF document. "
                    f"Extract ALL data and return JSON matching "
                    f"the target schema.\n\n"
                    f"{snapshot}"
                ),
            },
        ]
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    # Call Claude with retry on rate limits
    import time

    max_retries = 3
    for attempt in range(max_retries):
        try:
            response = anthropic_client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=4096,
                system=_PARSE_SYSTEM_PROMPT,
                messages=[{"role": "user", "content": content}],
            )
            break
        except Exception as exc:
            if "rate_limit" in str(exc).lower() and attempt < max_retries - 1:
                wait = 15 * (attempt + 1)
                logger.warning("Rate limited for %s, retrying in %ds...", filename, wait)
                time.sleep(wait)
                continue
            logger.exception("Claude call failed in parse_file_with_claude for %s", filename)
            raise RuntimeError(f"Claude API call failed: {exc}") from exc

    # Extract text from response
    text = ""
    for block in response.content:
        if block.type == "text":
            text += block.text

    # Strip markdown fences if Claude added them despite instructions
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*\n?", "", text)
        text = re.sub(r"\n?```\s*$", "", text)
        text = text.strip()

    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        # Attempt to repair truncated JSON (e.g. max_tokens hit mid-array)
        data = _repair_truncated_json(text)
        if data is None:
            logger.error("Claude returned invalid JSON for %s: %s", filename, text[:500])
            raise ValueError(f"Claude returned unparseable JSON for {filename}")

    # Ensure all expected top-level keys exist with sensible defaults
    result: dict[str, Any] = {
        "performance": data.get("performance", []),
        "allocation": data.get("allocation", []),
        "roll_forward": data.get("roll_forward", []),
        "contributions": data.get("contributions", []),
        "mv_history": data.get("mv_history", []),
        "sma_holdings": data.get("sma_holdings", []),
        "sma_summary": data.get("sma_summary", {}),
    }

    # If Claude returned a summary but no holdings, extract holdings deterministically
    if ext in (".xlsx", ".xls") and result.get("sma_summary") and not result.get("sma_holdings"):
        holdings = _extract_holdings_from_snapshot(file_path)
        if holdings:
            result["sma_holdings"] = holdings
            # Update summary num_holdings if it was 0
            if isinstance(result["sma_summary"], dict):
                result["sma_summary"]["num_holdings"] = len(holdings)

    return result
